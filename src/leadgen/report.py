from __future__ import annotations

import html
import logging
from datetime import UTC, datetime

import pandas as pd

from leadgen.config import Settings
from leadgen.notify import send_telegram_file
from leadgen.storage import LeadStore

logger = logging.getLogger("leadgen.report")

DRAFT_PROMPT = """Write a short (under 60 words), non-salesy outreach note
for a freelance GenAI/AI-automation engineer to send in reply to this lead.
Reference their specific need, don't be generic, no exclamation marks,
no "I hope this finds you well" filler.

Lead title: {title}
Lead snippet: {snippet}

Return ONLY the note text, nothing else.
"""


def draft_note(client, model: str, title: str, snippet: str) -> str:
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": DRAFT_PROMPT.format(title=title, snippet=snippet)}],
        )
        return resp.choices[0].message.content
    except Exception as e:  # noqa: BLE001 - a failed draft shouldn't kill the whole report
        logger.warning("draft generation failed: %s", e)
        return "[draft unavailable — write manually]"


def _get_draft_client_and_model(settings: Settings):
    """Priority: Grok/xAI (already used for search, one less key to manage)
    -> Groq (free tier is genuinely generous: 14,400 req/day, no card, but
    NOT the same product as Grok/xAI — no web/X search, drafting only)
    -> OpenRouter (fallback; free models there are less consistently
    available). Returns (None, None) if none configured — drafting skipped."""
    from openai import OpenAI

    if settings.xai_api_key:
        return OpenAI(api_key=settings.xai_api_key, base_url="https://api.x.ai/v1"), settings.xai_model
    if settings.groq_api_key:
        return (
            OpenAI(api_key=settings.groq_api_key, base_url="https://api.groq.com/openai/v1"),
            settings.groq_model,
        )
    if settings.openrouter_api_key:
        headers = {
            "HTTP-Referer": "https://github.com/Shweta-Mishra-ai/leadgen",
            "X-Title": "leadgen",
        }
        return (
            OpenAI(
                api_key=settings.openrouter_api_key,
                base_url="https://openrouter.ai/api/v1",
                default_headers=headers,
            ),
            settings.openrouter_model,
        )
    return None, None


def _apply_excel_formatting(filepath: str) -> None:
    """Applies professional formatting, custom column widths, dark header styling,
    and clickable Excel hyperlinks to the generated .xlsx report."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(filepath)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="2563EB", underline="single")

    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 28

        # Format header
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Format data rows
        url_col_idx = None
        for col in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=1, column=col).value or "").lower()
            if "url" in val:
                url_col_idx = col

        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 35
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if col == url_col_idx and cell.value and str(cell.value).startswith("http"):
                    raw_url = str(cell.value)
                    cell.value = f'=HYPERLINK("{raw_url}", "Open Link")'
                    cell.font = link_font
                else:
                    cell.font = data_font

        # Auto-adjust column widths
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            col_name = str(sheet.cell(row=1, column=col[0].column).value or "").lower()

            if "title" in col_name:
                sheet.column_dimensions[col_letter].width = 45
            elif "snippet" in col_name or "draft" in col_name:
                sheet.column_dimensions[col_letter].width = 65
            elif "url" in col_name:
                sheet.column_dimensions[col_letter].width = 16
            elif "score" in col_name:
                sheet.column_dimensions[col_letter].width = 12
            else:
                sheet.column_dimensions[col_letter].width = 20

    wb.save(filepath)


def generate_report(settings: Settings, store: LeadStore) -> str:
    from leadgen.notify import send_telegram_message

    all_leads = store.all_leads()
    if not all_leads:
        logger.warning("no leads in store, nothing to report")
        raise ValueError("No leads found — run the pipeline first")

    # Clean & Select Lead Columns
    clean_leads = []
    for row in all_leads:
        clean_leads.append({
            "Score": row.get("score", 0),
            "Source": row.get("source", ""),
            "Title": row.get("title", ""),
            "URL": row.get("url", ""),
            "Snippet": row.get("snippet", ""),
            "Created": row.get("created", ""),
            "Found At": row.get("found_at", ""),
        })

    df = pd.DataFrame(clean_leads)

    drafts = []
    client, model = _get_draft_client_and_model(settings)
    if client:
        top = store.top_leads(limit=settings.top_n_for_drafts)
        for lead in top:
            note = draft_note(client, model, lead["title"], lead["snippet"])
            drafts.append({
                "Score": lead.get("score", 0),
                "Source": lead.get("source", ""),
                "Title": lead.get("title", ""),
                "URL": lead.get("url", ""),
                "Outreach Draft Note": note,
            })
    else:
        logger.info("No drafting keys configured (XAI, GROQ, or OPENROUTER), skipping draft generation")

    drafts_df = pd.DataFrame(drafts) if drafts else pd.DataFrame(
        columns=["Score", "Source", "Title", "URL", "Outreach Draft Note"]
    )

    today = datetime.now(UTC).strftime("%Y-%m-%d")
    report_path = f"leads_report_{today}.xlsx"

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        drafts_df.to_excel(writer, sheet_name="Drafts", index=False)

    # Format Excel with openpyxl styling & clickable hyperlinks
    try:
        _apply_excel_formatting(report_path)
    except Exception as e:  # noqa: BLE001
        logger.warning("Excel styling application warning: %s", e)

    # Build Rich Telegram Summary Message
    top_5 = clean_leads[:5]
    summary_msg = [
        f"📊 <b>Daily Lead Report — {today}</b>",
        f"<b>Total Leads:</b> {len(clean_leads)} | <b>Drafts Ready:</b> {len(drafts_df)}",
        "",
        "<b>🌟 Top 5 Scored Leads:</b>",
    ]
    for idx, lead in enumerate(top_5, 1):
        title = html.escape(str(lead['Title']))
        summary_msg.append(
            f"{idx}. <b>{title}</b>\n"
            f"   ⭐ Score: <code>{lead['Score']}</code> | Source: <i>{lead['Source']}</i>\n"
            f"   🔗 <a href=\"{lead['URL']}\">View Lead Link</a>\n"
        )
    summary_msg.append("📎 <i>Complete formatted Excel report attached below.</i>")
    formatted_text = "\n".join(summary_msg)

    # Send Telegram Message + Excel Document
    send_telegram_message(
        settings.telegram_bot_token, settings.telegram_chat_id, formatted_text, parse_mode="HTML"
    )
    send_telegram_file(
        settings.telegram_bot_token, settings.telegram_chat_id, report_path,
        caption=f"📁 <b>leads_report_{today}.xlsx</b> ({len(clean_leads)} total leads)",
        parse_mode="HTML",
    )
    return report_path
